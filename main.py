import os
import glob
import openpyxl
import pyodbc
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import argparse

def get_excel_files(fileshare_path):
    """Returns a list of Excel files in the specified directory and its subdirectories."""
    excel_files = []
    for root, dirs, files in os.walk(fileshare_path):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                excel_files.append(file_path)
    return excel_files

def get_keywords(keywords_file_path):
    """Returns a list of keywords read from the specified file."""
    with open(keywords_file_path, 'r') as keywords_file:
        keywords = keywords_file.read().splitlines()
    return keywords

def get_workbook_metadata(workbook):
    """Returns the author and last modified date of the specified workbook."""
    author = workbook.properties.creator
    last_modified = workbook.properties.modified
    return author, last_modified

def write_results_to_excel(results_worksheet, row_index, file_path, author, last_modified, formula):
    """Writes the results for a single Excel file to the specified worksheet."""
    results_worksheet.cell(row=row_index, column=1, value=os.path.basename(file_path))
    results_worksheet.cell(row=row_index, column=2, value=file_path)
    results_worksheet.cell(row=row_index, column=3, value=author)
    results_worksheet.cell(row=row_index, column=4, value=last_modified)
    results_worksheet.cell(row=row_index, column=5, value=formula)
    results_worksheet.cell(row=row_index, column=6, value="")
    row_index += 1
    return row_index

def write_results_to_database(cursor, file_path, author, last_modified, formula):
    """Writes the results for a single Excel file to the specified database."""
    cursor.execute("INSERT INTO EUC_Check (Filename, Path, Author, LastModified, Formula, Keywords) VALUES (?, ?, ?, ?, ?, ?)",
                   os.path.basename(file_path), file_path, author, last_modified, formula, "")

def scan_excel_files(fileshare_path, keywords, results_worksheet, cursor):
    """Scans all Excel files in the specified directory and its subdirectories for formulas or keywords."""
    row_index = 2
    for file_path in get_excel_files(fileshare_path):
        try:
            workbook = load_workbook(file_path, data_only=True)
            for sheet in workbook.worksheets:
                if sheet.formulae:
                    author, last_modified = get_workbook_metadata(workbook)
                    for formula in sheet.formulae:
                        row_index = write_results_to_excel(results_worksheet, row_index, file_path, author, last_modified, formula)
                        write_results_to_database(cursor, file_path, author, last_modified, formula)
                    break
            else:
                with open(file_path, 'r') as file:
                    contents = file.read()
                    if any(keyword in contents for keyword in keywords):
                        author, last_modified = get_workbook_metadata(workbook)
                        row_index = write_results_to_excel(results_worksheet, row_index, file_path, author, last_modified, "")
                        write_results_to_database(cursor, file_path, author, last_modified, "")
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
    return results_worksheet

if __name__ == '__main__':
    # create an argument parser to accept
